#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ExcelFormatter - Classe para criar arquivos Excel formatados com dados do Zap Imóveis
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import warnings

warnings.filterwarnings("ignore")

class ExcelFormatter:
    def __init__(self):
        self.workbook = Workbook()
        self.worksheet_names = {
            'dados_completos': 'Dados Completos',
            'resumo_estatistico': 'Resumo Estatístico',
            'analise_precos': 'Análise de Preços',
            'analise_areas': 'Análise de Áreas',
            'top_imoveis': 'Top Imóveis',
            'filtros_especiais': 'Filtros Especiais'
        }
    
    def criar_cabecalho_formatado(self, worksheet, titulo, colunas):
        """Cria cabeçalho formatado para cada aba"""
        # Mesclar células para o título
        worksheet.merge_cells(f'A1:{chr(65 + len(colunas) - 1)}1')
        worksheet['A1'] = titulo
        worksheet['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        worksheet['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Configurar cabeçalhos das colunas (começando da coluna 1, não 2)
        for idx, coluna in enumerate(colunas, 1):
            cell = worksheet.cell(row=2, column=idx, value=coluna)
            cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adicionar bordas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
    
    def aplicar_formato_dados(self, worksheet, start_row, colunas):
        """Aplica formatação aos dados"""
        # Formatar células de dados
        for row in range(start_row, worksheet.max_row + 1):
            for col in range(1, len(colunas) + 1):
                cell = worksheet.cell(row=row, column=col)  # Corrigido: col em vez de col + 1
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Formatação específica para colunas numéricas
                if col > 2:  # Colunas numéricas (M2, Preço, R$/M2, etc.)
                    if 'Preco' in colunas[col-1] or 'Condominio' in colunas[col-1] or 'IPTU' in colunas[col-1]:
                        cell.number_format = 'R$ #,##0.00'
                    elif 'M2' in colunas[col-1]:
                        cell.number_format = '#,##0.00'
                    elif 'R$/M2' in colunas[col-1]:
                        cell.number_format = 'R$ #,##0.00'
                
                # Adicionar bordas
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
                
                # Alternar cores das linhas
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    def ajustar_largura_colunas(self, worksheet, colunas):
        """Ajusta a largura das colunas automaticamente"""
        column_widths = {
            'Descrição': 60,
            'Endereco': 30,
            'M2': 10,
            'Quartos': 8,
            'Banheiros': 10,
            'Vagas': 8,
            'Preco': 15,
            'Condominio': 15,
            'IPTU': 15,
            'R$/M2': 15
        }
        
        for idx, coluna in enumerate(colunas, 1):
            width = column_widths.get(coluna, 15)
            worksheet.column_dimensions[chr(64 + idx)].width = width  # Corrigido: idx em vez de idx + 1
    
    def criar_aba_dados_completos(self, df):
        """Cria aba com todos os dados"""
        worksheet = self.workbook.active
        worksheet.title = self.worksheet_names['dados_completos']
        
        colunas = df.columns.tolist()
        self.criar_cabecalho_formatado(worksheet, 'DADOS COMPLETOS - IMÓVEIS ZAP', colunas)
        
        # Adicionar dados
        for r in dataframe_to_rows(df, index=False, header=False):
            worksheet.append(r)
        
        self.aplicar_formato_dados(worksheet, 3, colunas)
        self.ajustar_largura_colunas(worksheet, colunas)
    
    def criar_aba_resumo_estatistico(self, df):
        """Cria aba com resumo estatístico"""
        worksheet = self.workbook.create_sheet(self.worksheet_names['resumo_estatistico'])
        
        # Calcular estatísticas
        stats = self.calcular_estatisticas_completas(df)
        
        # Criar cabeçalho
        colunas = ['Métrica', 'Valor']
        self.criar_cabecalho_formatado(worksheet, 'RESUMO ESTATÍSTICO', colunas)
        
        # Adicionar estatísticas
        metricas = [
            ('Total de Imóveis', stats['total_imoveis']),
            ('Preço Médio', f"R$ {stats['preco_medio']:,.2f}"),
            ('Preço Mínimo', f"R$ {stats['preco_minimo']:,.2f}"),
            ('Preço Máximo', f"R$ {stats['preco_maximo']:,.2f}"),
            ('Área Média', f"{stats['area_media']:.2f} m²"),
            ('Área Mínima', f"{stats['area_minima']:.2f} m²"),
            ('Área Máxima', f"{stats['area_maxima']:.2f} m²"),
            ('Preço por m² Médio', f"R$ {stats['preco_por_m2_medio']:,.2f}"),
            ('Preço por m² Mínimo', f"R$ {stats['preco_por_m2_minimo']:,.2f}"),
            ('Preço por m² Máximo', f"R$ {stats['preco_por_m2_maximo']:,.2f}"),
            ('Mediana Preço por m²', f"R$ {stats['mediana_preco_por_m2']:,.2f}"),
            ('Desvio Padrão Preço por m²', f"R$ {stats['desvio_preco_por_m2']:,.2f}"),
            ('Coeficiente de Variação', f"{stats['coef_variacao']:.4f}"),
            ('Quartos Médios', f"{stats['quartos_medio']:.1f}"),
            ('Banheiros Médios', f"{stats['banheiros_medio']:.1f}"),
            ('Vagas Médias', f"{stats['vagas_medio']:.1f}")
        ]
        
        for metrica, valor in metricas:
            worksheet.append([metrica, valor])
        
        self.aplicar_formato_dados(worksheet, 3, colunas)
        self.ajustar_largura_colunas(worksheet, colunas)
    
    def criar_aba_analise_precos(self, df):
        """Cria aba com análise de preços"""
        worksheet = self.workbook.create_sheet(self.worksheet_names['analise_precos'])
        
        # Criar faixas de preço
        df_precos = df.copy()
        df_precos['Faixa_Preco'] = pd.cut(df_precos['Preco'], 
                                        bins=[0, 100000, 200000, 300000, 400000, 500000, float('inf')],
                                        labels=['Até R$ 100k', 'R$ 100k-200k', 'R$ 200k-300k', 
                                               'R$ 300k-400k', 'R$ 400k-500k', 'Acima de R$ 500k'])
        
        # Agrupar por faixa de preço
        analise_precos = df_precos.groupby('Faixa_Preco').agg({
            'Preco': ['count', 'mean', 'min', 'max'],
            'M2': 'mean',
            'R$/M2': 'mean'
        }).round(2)
        
        # Flatten das colunas
        analise_precos.columns = ['Quantidade', 'Preço_Médio', 'Preço_Mínimo', 'Preço_Máximo', 
                                'Área_Média', 'Preço_por_m2_Médio']
        analise_precos = analise_precos.reset_index()
        
        colunas = analise_precos.columns.tolist()
        self.criar_cabecalho_formatado(worksheet, 'ANÁLISE POR FAIXAS DE PREÇO', colunas)
        
        # Adicionar dados
        for r in dataframe_to_rows(analise_precos, index=False, header=False):
            worksheet.append(r)
        
        self.aplicar_formato_dados(worksheet, 3, colunas)
        self.ajustar_largura_colunas(worksheet, colunas)
    
    def criar_aba_analise_areas(self, df):
        """Cria aba com análise de áreas"""
        worksheet = self.workbook.create_sheet(self.worksheet_names['analise_areas'])
        
        # Criar faixas de área
        df_areas = df.copy()
        df_areas['Faixa_Area'] = pd.cut(df_areas['M2'], 
                                       bins=[0, 40, 60, 80, 100, 120, float('inf')],
                                       labels=['Até 40m²', '40-60m²', '60-80m²', 
                                              '80-100m²', '100-120m²', 'Acima de 120m²'])
        
        # Agrupar por faixa de área
        analise_areas = df_areas.groupby('Faixa_Area').agg({
            'M2': ['count', 'mean', 'min', 'max'],
            'Preco': 'mean',
            'R$/M2': 'mean'
        }).round(2)
        
        # Flatten das colunas
        analise_areas.columns = ['Quantidade', 'Área_Média', 'Área_Mínima', 'Área_Máxima', 
                               'Preço_Médio', 'Preço_por_m2_Médio']
        analise_areas = analise_areas.reset_index()
        
        colunas = analise_areas.columns.tolist()
        self.criar_cabecalho_formatado(worksheet, 'ANÁLISE POR FAIXAS DE ÁREA', colunas)
        
        # Adicionar dados
        for r in dataframe_to_rows(analise_areas, index=False, header=False):
            worksheet.append(r)
        
        self.aplicar_formato_dados(worksheet, 3, colunas)
        self.ajustar_largura_colunas(worksheet, colunas)
    
    def criar_aba_top_imoveis(self, df):
        """Cria aba com os melhores imóveis por diferentes critérios"""
        worksheet = self.workbook.create_sheet(self.worksheet_names['top_imoveis'])
        
        # Melhores imóveis por preço por m² (menor valor)
        df_ordenado = df.sort_values('R$/M2').head(10)
        
        colunas = df_ordenado.columns.tolist()
        self.criar_cabecalho_formatado(worksheet, 'TOP 10 - MELHORES PREÇOS POR M²', colunas)
        
        # Adicionar dados
        for r in dataframe_to_rows(df_ordenado, index=False, header=False):
            worksheet.append(r)
        
        self.aplicar_formato_dados(worksheet, 3, colunas)
        self.ajustar_largura_colunas(worksheet, colunas)
    
    def criar_aba_filtros_especiais(self, df):
        """Cria aba com filtros especiais"""
        worksheet = self.workbook.create_sheet(self.worksheet_names['filtros_especiais'])
        
        # Filtros especiais
        filtros = []
        
        # Imóveis com 3 quartos e 2 banheiros
        if 'Quartos' in df.columns and 'Banheiros' in df.columns:
            filtro_3q2b = df[(df['Quartos'] == 3) & (df['Banheiros'] == 2)]
            if not filtro_3q2b.empty:
                filtros.append(('Imóveis 3 Quartos + 2 Banheiros', filtro_3q2b))
        
        # Imóveis com vaga de garagem
        if 'Vagas' in df.columns:
            filtro_com_vaga = df[df['Vagas'] > 0]
            if not filtro_com_vaga.empty:
                filtros.append(('Imóveis com Vaga de Garagem', filtro_com_vaga))
        
        # Imóveis até R$ 200.000
        filtro_ate_200k = df[df['Preco'] <= 200000]
        if not filtro_ate_200k.empty:
            filtros.append(('Imóveis até R$ 200.000', filtro_ate_200k))
        
        # Imóveis com área entre 50-80m²
        filtro_area_50_80 = df[(df['M2'] >= 50) & (df['M2'] <= 80)]
        if not filtro_area_50_80.empty:
            filtros.append(('Imóveis 50-80m²', filtro_area_50_80))
        
        # Adicionar cada filtro como uma seção
        row_atual = 1
        for nome_filtro, dados_filtro in filtros:
            # Título do filtro
            worksheet.merge_cells(f'A{row_atual}:{chr(65 + len(dados_filtro.columns) - 1)}{row_atual}')
            worksheet[f'A{row_atual}'] = nome_filtro
            worksheet[f'A{row_atual}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            worksheet[f'A{row_atual}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            worksheet[f'A{row_atual}'].alignment = Alignment(horizontal='center', vertical='center')
            
            row_atual += 1
            
            # Cabeçalhos das colunas
            colunas = dados_filtro.columns.tolist()
            for idx, coluna in enumerate(colunas):
                cell = worksheet.cell(row=row_atual, column=idx + 1, value=coluna)
                cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row_atual += 1
            
            # Dados
            for _, row in dados_filtro.iterrows():
                for idx, valor in enumerate(row):
                    worksheet.cell(row=row_atual, column=idx + 1, value=valor)
                row_atual += 1
            
            row_atual += 2  # Espaço entre filtros
        
        # Aplicar formatação geral
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value and not cell.font.bold:
                    cell.font = Font(name='Arial', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        self.ajustar_largura_colunas(worksheet, df.columns.tolist())
    
    def calcular_estatisticas_completas(self, df):
        """Calcula estatísticas completas dos dados"""
        stats = {
            'total_imoveis': len(df),
            'preco_medio': df['Preco'].mean(),
            'preco_minimo': df['Preco'].min(),
            'preco_maximo': df['Preco'].max(),
            'area_media': df['M2'].mean(),
            'area_minima': df['M2'].min(),
            'area_maxima': df['M2'].max(),
            'preco_por_m2_medio': df['R$/M2'].mean(),
            'preco_por_m2_minimo': df['R$/M2'].min(),
            'preco_por_m2_maximo': df['R$/M2'].max(),
            'mediana_preco_por_m2': df['R$/M2'].median(),
            'desvio_preco_por_m2': df['R$/M2'].std(),
            'coef_variacao': df['R$/M2'].std() / df['R$/M2'].mean(),
            'quartos_medio': df['Quartos'].mean() if 'Quartos' in df.columns else 0,
            'banheiros_medio': df['Banheiros'].mean() if 'Banheiros' in df.columns else 0,
            'vagas_medio': df['Vagas'].mean() if 'Vagas' in df.columns else 0
        }
        return stats
    
    def gerar_excel_formatado(self, df, nome_arquivo=None):
        """Gera arquivo Excel formatado com todas as abas"""
        if nome_arquivo is None:
            # Criar pasta arquivos se não existir
            pasta_arquivos = "arquivos"
            if not os.path.exists(pasta_arquivos):
                os.makedirs(pasta_arquivos)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = os.path.join(pasta_arquivos, f'dados_zap_formatado_{timestamp}.xlsx')
        
        try:
            # Remover coluna URL se existir
            if 'URL' in df.columns:
                df = df.drop('URL', axis=1)
                print("🗑️ Coluna 'URL' removida do Excel")
            
            # Criar todas as abas
            self.criar_aba_dados_completos(df)
            self.criar_aba_resumo_estatistico(df)
            self.criar_aba_analise_precos(df)
            self.criar_aba_analise_areas(df)
            self.criar_aba_top_imoveis(df)
            self.criar_aba_filtros_especiais(df)
            
            # Salvar arquivo
            self.workbook.save(nome_arquivo)
            print(f"✅ Arquivo Excel formatado criado com sucesso: {nome_arquivo}")
            print(f"📊 Total de abas criadas: {len(self.workbook.worksheets)}")
            print(f"📈 Total de imóveis processados: {len(df)}")
            
            return nome_arquivo
            
        except Exception as e:
            print(f"❌ Erro ao criar arquivo Excel: {e}")
            return None

def processar_csv_para_excel(arquivo_csv):
    """Processa um arquivo CSV existente e gera Excel formatado"""
    try:
        # Ler arquivo CSV
        df = pd.read_csv(arquivo_csv)
        print(f"📁 Arquivo CSV carregado: {arquivo_csv}")
        print(f"📊 Total de registros: {len(df)}")
        
        # Criar formatter e gerar Excel
        formatter = ExcelFormatter()
        nome_excel = formatter.gerar_excel_formatado(df)
        
        return nome_excel
        
    except Exception as e:
        print(f"❌ Erro ao processar arquivo CSV: {e}")
        return None

if __name__ == "__main__":
    # Exemplo de uso
    arquivos_csv = [f for f in os.listdir('.') if f.startswith('dados_zap_') and f.endswith('.csv')]
    
    if arquivos_csv:
        print("📋 Arquivos CSV encontrados:")
        for i, arquivo in enumerate(arquivos_csv, 1):
            print(f"  {i}. {arquivo}")
        
        escolha = input("\nDigite o número do arquivo para processar (ou Enter para o mais recente): ")
        
        if escolha.isdigit() and 1 <= int(escolha) <= len(arquivos_csv):
            arquivo_selecionado = arquivos_csv[int(escolha) - 1]
        else:
            arquivo_selecionado = arquivos_csv[-1]  # Mais recente
        
        print(f"\n🔄 Processando: {arquivo_selecionado}")
        excel_gerado = processar_csv_para_excel(arquivo_selecionado)
        
        if excel_gerado:
            print(f"\n🎉 Processo concluído! Arquivo Excel criado: {excel_gerado}")
        else:
            print("\n❌ Falha ao criar arquivo Excel")
    else:
        print("❌ Nenhum arquivo CSV encontrado na pasta atual")